from openpyxl import *
from tkinter import *
  
# globally declare wb and sheet variable 
  
# opening the existing excel file 
wb = load_workbook('C:\\Users\\Administrator\\Desktop\\excel.xlsx')
  
# create the sheet object 
sheet = wb.active 
  
  
def excel(): 
      
    # resize the width of columns in 
    # excel spreadsheet 
    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 12
    sheet.column_dimensions['D'].width = 30
    sheet.column_dimensions['E'].width = 12
    sheet.column_dimensions['F'].width = 12
    sheet.column_dimensions['G'].width = 12
  
    # write given data to an excel spreadsheet 
    # at particular location 
    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Surname"
    sheet.cell(row=1, column=3).value = "Id"
    sheet.cell(row=1, column=4).value = "Email"
    sheet.cell(row=1, column=5).value = "Midterm1"
    sheet.cell(row=1, column=6).value = "Midterm2"
    sheet.cell(row=1, column=7).value = "Final"
  
  
# Function to set focus (cursor)
def focus1(event):
    # set focus on the course_field box
    name_field.focus_set()


# Function to set focus
def focus2(event):
    # set focus on the sem_field box
    surname_field.focus_set()


# Function to set focus
def focus3(event):
    # set focus on the form_no_field box
    id_field.focus_set()


# Function to set focus
def focus4(event):
    # set focus on the contact_no_field box
    email_field.focus_set()


# Function to set focus
def focus5(event):
    # set focus on the email_id_field box
    midterm1_field.focus_set()


# Function to set focus
def focus6(event):
    # set focus on the address_field box
    midterm2.focus_set()

def focus7(event):
    final_field.focus_set()


# Function for clearing the 
# contents of text entry boxes 
def clear(): 

    name_field.delete(0, END) 
    surname_field.delete(0, END)
    id_field.delete(0, END)
    email_field.delete(0, END)
    midterm1_field.delete(0, END)
    midterm2_field.delete(0, END)
    final_field.delete(0, END)
  
  
# Function to take data from GUI  
# window and write to an excel file 
def insert():
    current_row = sheet.max_row
    sheet.cell(row=current_row + 1, column=1).value = name_field.get()
    sheet.cell(row=current_row + 1, column=2).value = surname_field.get()
    sheet.cell(row=current_row + 1, column=3).value = id_field.get()
    sheet.cell(row=current_row + 1, column=4).value = email_field.get()
    sheet.cell(row=current_row + 1, column=5).value = midterm1_field.get()
    sheet.cell(row=current_row + 1, column=6).value = midterm2_field.get()
    sheet.cell(row=current_row + 1, column=7).value = final_field.get()

    # save the file
    wb.save('C:\\Users\\Administrator\\Desktop\\excel.xlsx')

    # call the clear() function
    clear()


  
  
# Driver code 
if __name__ == "__main__": 
      
    # create a GUI window 
    root = Tk() 
  
    # set the background colour of GUI window 
    root.configure(background='light green') 
  
    # set the title of GUI window 
    root.title("Grade Assistant")
  
    # set the configuration of GUI window 
    root.geometry("500x300") 
  
    excel() 
  
    # create a Form label 
    heading = Label(root, text="Form", bg="light green") 
  
    # create a Name label 
    name = Label(root, text="Name", bg="light green") 
  
    # create a Course label 
    surname = Label(root, text="Surname", bg="light green")
  
    # create a Semester label 
    id = Label(root, text="Id", bg="light green")
  
    # create a Form No. lable 
    email = Label(root, text="Email", bg="light green")
  
    # create a Contact No. label 
    midterm1 = Label(root, text="Midterm1", bg="light green")
  
    # create a Email id label 
    midterm2 = Label(root, text="Midterm2", bg="light green")
  
    # create a address label 
    final = Label(root, text="Final", bg="light green")
  
    # grid method is used for placing 
    # the widgets at respective positions 
    # in table like structure . 
    heading.grid(row=0, column=1) 
    name.grid(row=1, column=0) 
    surname.grid(row=2, column=0)
    id.grid(row=3, column=0)
    email.grid(row=4, column=0)
    midterm1.grid(row=5, column=0)
    midterm2.grid(row=6, column=0)
    final.grid(row=7, column=0)
  
    # create a text entry box 
    # for typing the information 
    name_field = Entry(root) 
    surname_field = Entry(root)
    id_field = Entry(root)
    email_field = Entry(root)
    midterm1_field = Entry(root)
    midterm2_field = Entry(root)
    final_field = Entry(root)
  
    # bind method of widget is used for 
    # the binding the function with the events 
  
    # whenever the enter key is pressed 
    # then call the focus1 function 
    name_field.bind("<Return>", focus1)

    # whenever the enter key is pressed
    # then call the focus2 function
    surname_field.bind("<Return>", focus2)

    # whenever the enter key is pressed
    # then call the focus3 function
    id_field.bind("<Return>", focus3)

    # whenever the enter key is pressed
    # then call the focus4 function
    email_field.bind("<Return>", focus4)

    # whenever the enter key is pressed
    # then call the focus5 function
    midterm1_field.bind("<Return>", focus5)

    # whenever the enter key is pressed
    # then call the focus6 function
    midterm2_field.bind("<Return>", focus6)

    final_field.bind("<Return>", focus7)
  
    # grid method is used for placing 
    # the widgets at respective positions 
    # in table like structure . 
    name_field.grid(row=1, column=1, ipadx="50")
    surname_field.grid(row=2, column=1, ipadx="50")
    id_field.grid(row=3, column=1, ipadx="50")
    email_field.grid(row=4, column=1, ipadx="50")
    midterm1_field.grid(row=5, column=1, ipadx="50")
    midterm2_field.grid(row=6, column=1, ipadx="50")
    final_field.grid(row=7, column=1, ipadx="50")
  
    # call excel function 
    excel() 
  
    # create a Submit Button and place into the root window 
    submit = Button(root, text="Submit", fg="Black", 
                            bg="Red", command=insert) 
    submit.grid(row=8, column=1) 
  
    # start the GUI 
    root.mainloop()